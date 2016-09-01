import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from peluqueria.models import Cliente, Caja

from datetime import datetime

def ArmoExcelCumple(cumple_mes):
    wb = openpyxl.Workbook()
    wb = Workbook(encoding='utf-8')
    wb.get_sheet_names()
    sheet = wb.active
    sheet.title = 'Cumpleaños'
    sheet['A1']='Apellido y Nombre'
    sheet['B1']='E-Mail'
    sheet['C1']='Teléfono'
    sheet['D1']='Cumpleaños'

    font_title = Font('name=Times New Roman', bold=True)

    sheet['A1'].font= font_title
    sheet['B1'].font= font_title
    sheet['C1'].font= font_title
    sheet['D1'].font= font_title

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20

    sheet.freeze_panes = 'A2'
    i = 1
    for cumple in cumple_mes:
         i = i + 1
         sheet['A'+ str(i)].value = cumple.Apellido+", "+cumple.Nombre
         sheet['B'+ str(i)].value = cumple.email
         sheet['C'+ str(i)].value = cumple.Teléfono
         sheet['D'+ str(i)].value = cumple.Cumpleaños.strftime("%d-%m")

    return save_virtual_workbook(wb)

def ArmoExcelCaja(caja_dia):
    wb = openpyxl.Workbook()
    wb = Workbook(encoding='utf-8')
    wb.get_sheet_names()
    sheet = wb.active
    sheet.title = 'Movimiento de Caja'
    sheet['A1']='Fecha Operacion'
    sheet['B1']='Tipo'
    sheet['C1']='Importe'
    sheet['D1']='Concepto'

    font_title = Font('name=Times New Roman', bold=True)

    sheet['A1'].font= font_title
    sheet['B1'].font= font_title
    sheet['C1'].font= font_title
    sheet['D1'].font= font_title

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20

    sheet.freeze_panes = 'A2'
    i = 1
    for caja in caja_dia:
         i = i + 1
         sheet['A'+ str(i)].value = caja.fecha_operacion.strftime("%d/%m/%Y")
         sheet['B'+ str(i)].value = caja.tipo_operacion.tipo_operacion
         sheet['C'+ str(i)].value = caja.importe if caja.tipo_operacion.tipo_operacion=='Ingreso' else -caja.importe
         sheet['D'+ str(i)].value = caja.concepto
    return save_virtual_workbook(wb)

def ImportarExcel(archivo):
    model = Cliente
    wb = load_workbook(filename = archivo, use_iterators=True)
    hoja = wb.get_sheet_names()[0]
    worksheet = wb.get_sheet_by_name(hoja)

    sheet = wb.active

    i = 1
    while True:
        i = i + 1
        if sheet['A'+ str(i)].value is None:
            break
        else:
            p = Cliente(
                Apellido = sheet['A'+ str(i)].value,
                Nombre = ("" if sheet['B'+ str(i)].value is None
                               else sheet['B'+ str(i)].value),
                Cumpleaños = (None if sheet['C'+ str(i)].value is None
                               else sheet['C'+ str(i)].value),
                email = ("" if sheet['D'+ str(i)].value is None
                               else sheet['D'+ str(i)].value),
                Teléfono = ("" if sheet['E'+ str(i)].value is None
                               else sheet['E'+ str(i)].value),
                Tratamiento = ("" if sheet['F'+ str(i)].value is None
                               else sheet['F'+ str(i)].value),
                fecha = sheet['G'+ str(i)].value
                )
            p.save()

def Fecha_a_anio_mes_dia(xfecha):
    wfecha = datetime.strptime(xfecha, "%d/%m/%Y").strftime("%Y-%m-%d")
    return wfecha
